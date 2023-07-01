import pandas as pd
import logging
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelLoader:
    SHEETS_TO_CHECK = ['How To', 'Bulkdatensammlung', 'SPA', 'SB', 'SD', 'KPIs', 'Dashboard', 'Auswertung']

    def __init__(self, file_path):
        self.file_path = file_path

    def load(self):
        try:
            logger.info(f'Loading Excel file from {self.file_path}')
            if not self.check_sheets():
                return None
            data = pd.read_excel(self.file_path, sheet_name='Bulkdatensammlung')
            logger.debug('Excel file loaded successfully.')
            return data
        except Exception as e:
            logger.error('An error occurred while loading the Excel file.', exc_info=True)
            raise e


    def save(self, data, output_path):
        try:
            logger.info(f'Saving data to {output_path}')
            
            # Load workbook
            book = load_workbook(self.file_path)
            
            # Remove the existing 'Bulkdatensammlung' sheet, if it exists
            if 'Bulkdatensammlung' in book.sheetnames:
                book.remove(book['Bulkdatensammlung'])
            
            # Save the workbook after removing the existing sheet
            book.save(output_path)

            # Create a Pandas Excel writer using openpyxl as the engine, in append mode
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
                # Write DataFrame to the 'Bulkdatensammlung' sheet in the workbook
                data.to_excel(writer, sheet_name='Bulkdatensammlung', index=False)

            logger.debug('Data saved successfully.')
        except Exception as e:
            logger.error('An error occurred while saving the data.', exc_info=True)
            raise e



    def check_sheets(self):
        try:
            logger.info('Starting sheet check.')
            workbook = load_workbook(filename=self.file_path, read_only=True)
            missing_sheets = [sheet for sheet in self.SHEETS_TO_CHECK if sheet not in workbook.sheetnames]
            if missing_sheets:
                logger.warning(f'Missing sheets in Excel file: {missing_sheets}')
                return False
            logger.info('Sheet check completed successfully.')
            return True
        except Exception as e:
            logger.error('An error occurred during the sheet check.', exc_info=True)
            raise e


class DataChecker:
    COLUMNS_TO_CHECK = ['Keyword Id (Read only)', 'Product Targeting Id (Read only)', 'Default Bid (nur bei SPA,SD)',
                        'Bid', 'Clicks', 'Spend', 'Sales', 'Datum des Downloads', 'Zeitraum (Tage)', 'KW-/Targeting-ID',
                        'Klick-Ind(Akt)', 'Ausgaben-Ind(Akt)', 'Umsatz-Ind(Akt)', 'Vergangene Tage bis Mitte des Zeitraums',
                        'Aktualitätsindex']

    def __init__(self, data):
        self.data = data

    def check_columns(self):
        try:
            logger.info('Starting data check.')
            missing_columns = [col for col in self.COLUMNS_TO_CHECK if col not in self.data.columns]
            if missing_columns:
                logger.warning(f'Missing columns in data: {missing_columns}')
                return False
            empty_columns = [col for col in self.data.columns if self.data[col].isna().all()]
            if empty_columns:
                logger.warning(f'Columns with no data: {empty_columns}')
                return False
            logger.info('Data check completed successfully.')
            return True
        except Exception as e:
            logger.error('An error occurred during the data check.', exc_info=True)
            raise e
